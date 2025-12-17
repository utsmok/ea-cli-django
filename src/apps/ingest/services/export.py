"""Excel export service for copyright items.

Phase A requirement:
- Create a folder per faculty
- Within each faculty folder create workflow-bucket workbooks:
  inbox.xlsx / in_progress.xlsx / done.xlsx / overview.xlsx
- Each workbook contains two sheets:
  "Complete data" (full raw export) + "Data entry" (editable fields + dropdowns)
- Create update_overview.csv at the root and update_info_*.txt per faculty

This is a Django-native implementation inspired by the legacy ea-cli exporter.
"""

from __future__ import annotations

import csv
import logging
import os
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import polars as pl
from django.conf import settings

from apps.core.models import CopyrightItem, Faculty, WorkflowStatus

from .excel_builder import ExcelBuilder

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class BucketStats:
    old: int
    new: int

    @property
    def delta(self) -> int:
        return self.new - self.old


class ExportService:
    """Exports copyright items into the legacy workflow folder structure."""

    BUCKETS: dict[str, set[str]] = {
        "inbox": {WorkflowStatus.TODO, "todo", "ToDo"},
        "in_progress": {
            WorkflowStatus.IN_PROGRESS,
            "InProgress",
            "in_progress",
            "inprogress",
        },
        "done": {WorkflowStatus.DONE, "Done", "done"},
    }

    def __init__(self, faculty_abbr: Optional[str] = None):
        self.faculty_abbr = faculty_abbr

    # ---------------------------------------------------------------------
    # Public API
    # ---------------------------------------------------------------------

    def export_workflow_tree(self, output_dir: Optional[Path] = None) -> dict[str, Any]:
        """Create the full faculty_sheets directory tree on disk."""
        output_dir = Path(
            output_dir
            or getattr(
                settings,
                "EXPORT_FACULTY_SHEETS_DIR",
                settings.PROJECT_ROOT / "exports" / "faculty_sheets",
            )
        )
        output_dir.mkdir(parents=True, exist_ok=True)

        faculties = self._get_faculty_codes()
        if self.faculty_abbr:
            faculties = [self.faculty_abbr]

        builder = ExcelBuilder()
        exported_files: list[Path] = []
        summary_rows: list[tuple[str, str, BucketStats]] = []

        for faculty in faculties:
            faculty_df = self._fetch_faculty_dataframe(faculty)
            if faculty_df.is_empty():
                logger.info("No items for faculty %s; skipping", faculty)
                continue

            faculty_dir = output_dir / faculty
            faculty_dir.mkdir(parents=True, exist_ok=True)

            buckets = self._bucketize(faculty_df)
            # overview is always all items
            buckets["overview"] = faculty_df

            stats_by_bucket: dict[str, BucketStats] = {}
            for bucket_name, bucket_df in buckets.items():
                if bucket_df.is_empty():
                    logger.info(
                        "No items for %s -> %s; skipping file creation",
                        faculty,
                        bucket_name,
                    )
                    stats_by_bucket[bucket_name] = BucketStats(
                        old=self._count_existing_rows(
                            faculty_dir / f"{bucket_name}.xlsx"
                        ),
                        new=0,
                    )
                    continue

                target_path = faculty_dir / f"{bucket_name}.xlsx"
                old_count = self._count_existing_rows(target_path)

                # Backup existing before overwriting
                if target_path.exists():
                    self._backup_existing_file(target_path, faculty_dir / "backups")

                wb = builder.build_workbook_for_dataframe(bucket_df)
                self._atomic_save_workbook(wb, target_path)

                exported_files.append(target_path)
                stats_by_bucket[bucket_name] = BucketStats(
                    old=old_count, new=bucket_df.height
                )
                summary_rows.append(
                    (faculty, bucket_name, stats_by_bucket[bucket_name])
                )

            self._write_update_info(faculty_dir, faculty, stats_by_bucket)

        self._append_update_overview_csv(output_dir, summary_rows)

        return {
            "output_dir": str(output_dir),
            "files": [str(p) for p in exported_files],
            "faculties": faculties,
        }

    # ---------------------------------------------------------------------
    # Data retrieval
    # ---------------------------------------------------------------------

    def _get_faculty_codes(self) -> list[str]:
        qs = Faculty.objects.all().values_list("abbreviation", flat=True)
        codes = sorted({c for c in qs if c})
        # Include UNM as a fallback bucket if present
        return codes

    def _fetch_faculty_dataframe(self, faculty: str) -> pl.DataFrame:
        """Fetch all items for a faculty as a Polars DataFrame in legacy column order."""
        cols = ExcelBuilder.FACULTY_SHEET_COLUMNS

        values = list(
            CopyrightItem.objects.filter(faculty__abbreviation=faculty).values(
                *[c for c in cols if c != "faculty"],
                "faculty__abbreviation",
            )
        )
        if not values:
            return pl.DataFrame()

        df = pl.DataFrame(values)
        if "faculty__abbreviation" in df.columns:
            df = df.rename({"faculty__abbreviation": "faculty"})

        # Ensure all columns exist
        for col in cols:
            if col not in df.columns:
                df = df.with_columns(pl.lit(None).alias(col))

        # Normalize workflow_status to canonical values
        df = df.with_columns(
            pl.col("workflow_status").fill_null(WorkflowStatus.TODO).cast(pl.Utf8)
        )

        return df.select(cols)

    # ---------------------------------------------------------------------
    # Bucketing & file I/O
    # ---------------------------------------------------------------------

    def _bucketize(self, df: pl.DataFrame) -> dict[str, pl.DataFrame]:
        wf = pl.col("workflow_status").fill_null(WorkflowStatus.TODO).cast(pl.Utf8)

        inbox_df = df.filter(
            wf.is_in(["ToDo", "todo", "TODO"])
            | (
                ~wf.is_in(
                    [
                        "ToDo",
                        "InProgress",
                        "Done",
                        "todo",
                        "done",
                        "inprogress",
                        "in_progress",
                    ]
                )
            )
        )
        in_progress_df = df.filter(
            wf.is_in(["InProgress", "inprogress", "in_progress"])
        )
        done_df = df.filter(wf.is_in(["Done", "done"]))

        return {
            "inbox": inbox_df,
            "in_progress": in_progress_df,
            "done": done_df,
        }

    def _atomic_save_workbook(self, wb, target_path: Path) -> None:
        target_path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = target_path.with_suffix(target_path.suffix + ".tmp")
        try:
            wb.save(tmp_path)
            os.replace(tmp_path, target_path)
        finally:
            try:
                if tmp_path.exists():
                    tmp_path.unlink(missing_ok=True)
            except Exception:
                # Best-effort cleanup
                pass

    def _count_existing_rows(self, path: Path) -> int:
        """Count existing exported rows based on the Complete data sheet (excluding header)."""
        if not path.exists():
            return 0
        try:
            import openpyxl

            wb = openpyxl.load_workbook(
                filename=str(path), read_only=True, data_only=True
            )
            if ExcelBuilder.COMPLETE_SHEET_NAME not in wb.sheetnames:
                return 0
            ws = wb[ExcelBuilder.COMPLETE_SHEET_NAME]
            # max_row includes header
            return max(0, (ws.max_row or 0) - 1)
        except Exception:
            return 0

    def _backup_existing_file(self, target_path: Path, backups_dir: Path) -> None:
        backups_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backups_dir / f"{target_path.stem}_{ts}{target_path.suffix}"
        try:
            target_path.replace(backup_path)
            logger.info("Backed up %s -> %s", target_path.name, backup_path.name)
        except Exception as e:
            logger.warning("Failed to backup %s: %s", target_path, e)

    def _append_update_overview_csv(
        self, output_dir: Path, rows: list[tuple[str, str, BucketStats]]
    ) -> None:
        if not rows:
            return

        summary_file = output_dir / "update_overview.csv"
        mode = "a" if summary_file.exists() else "w"
        with summary_file.open(mode, newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            if mode == "w":
                writer.writerow(
                    ["timestamp", "faculty", "bucket", "old", "new", "delta"]
                )

            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for faculty, bucket, stats in rows:
                if stats.delta == 0:
                    continue
                writer.writerow(
                    [now, faculty, bucket, stats.old, stats.new, stats.delta]
                )

    def _write_update_info(
        self, faculty_dir: Path, faculty: str, stats_by_bucket: dict[str, BucketStats]
    ) -> None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = faculty_dir / f"update_info_{ts}.txt"

        # Remove previous update_info files (keep only newest info, like the legacy does)
        for p in faculty_dir.glob("update_info_*.txt"):
            try:
                p.unlink(missing_ok=True)
            except Exception:
                pass

        def _line(text: str = "") -> str:
            return text + "\n"

        with path.open("w", encoding="utf-8") as fh:
            fh.write(_line("Update information for".center(40)))
            fh.write(_line(faculty.center(40)))
            fh.write(_line(""))
            fh.write(_line("Last sync with main database:".center(40)))
            fh.write(_line(datetime.now().strftime("%Y-%m-%d -- %H:%M:%S").center(40)))
            fh.write(_line(""))

            # Table
            fh.write(_line(("-" * 12 + "-" * 5 + "-" * 5 + "-" * 5).center(40)))
            fh.write(
                _line((f"{'Sheet':<12}|{'Old':^5}|{'New':^5}|{'Δ':^5}").center(40))
            )
            fh.write(
                _line(
                    ("-" * 12 + "+" + "-" * 5 + "+" + "-" * 5 + "+" + "-" * 5).center(
                        40
                    )
                )
            )
            for bucket_name in ["inbox", "in_progress", "done", "overview"]:
                stats = stats_by_bucket.get(bucket_name, BucketStats(old=0, new=0))
                fh.write(
                    _line(
                        (
                            f"{bucket_name:<12}|{stats.old:^5}|{stats.new:^5}|{stats.delta:^+5}"
                        ).center(40)
                    )
                )
            fh.write(_line(("-" * 12 + "-" * 5 + "-" * 5 + "-" * 5).center(40)))
