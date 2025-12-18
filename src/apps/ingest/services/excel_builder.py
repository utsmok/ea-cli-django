"""
Excel export service for backward compatibility.

Generates faculty-oriented Excel workbooks that mirror the legacy data-entry
layout and formatting. This is a port of `ea-cli/easy_access/sheets/sheet.py`.
"""
from __future__ import annotations

import hashlib
from io import BytesIO
from typing import Any, Iterable

import polars as pl
from loguru import logger
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table as ExcelTable
from openpyxl.worksheet.table import TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from apps.core.models import (
    ClassificationV2,
    CopyrightItem,
    Lengte,
    OvernameStatus,
    WorkflowStatus,
)

from . import export_config

ENUM_MAP = {
    "v2_manual_classification": ClassificationV2,
    "v2_overnamestatus": OvernameStatus,
    "v2_lengte": Lengte,
    "workflow_status": WorkflowStatus,
}

class ExcelBuilder:
    """
    Builds Excel workbooks for faculty data-entry/export.
    Port of the legacy `ea-cli` export routines.
    """

    COMPLETE_SHEET_NAME = "Complete data"
    DATA_ENTRY_SHEET_NAME = "Data entry"

    def build_workbook_for_dataframe(
        self, df: pl.DataFrame, style_iter: int = 9
    ) -> Workbook:
        """
        Build a legacy-compliant workbook for a single exported dataset.

        Creates two sheets:
        - "Complete data" (all columns, read-only)
        - "Data entry" (subset of columns, editable with dropdown validations)
        """
        wb = Workbook()

        # 1. Complete data sheet (read-only)
        ws_complete = wb.active
        ws_complete.title = self.COMPLETE_SHEET_NAME
        
        # Create ColumnConfig objects for the complete data sheet
        complete_data_configs = [export_config.ColumnConfig(name=col) for col in export_config.COMPLETE_DATA_COLUMN_ORDER]
        
        self._write_dataframe_to_sheet(
            ws_complete, df, complete_data_configs
        )
        self._protect_sheet_all_locked(ws_complete)

        # 2. Data entry sheet (editable)
        ws_entry = wb.create_sheet(title=self.DATA_ENTRY_SHEET_NAME)
        self._write_dataframe_to_sheet(
            ws_entry, df, export_config.DATA_ENTRY_COLUMNS
        )
        self._apply_styling_and_validation(ws_entry, df)
        # Create an Excel table object
        self._create_table(ws_entry, df.height, style_iter)

        # Make the Data entry sheet the active one
        wb.active = wb.sheetnames.index(self.DATA_ENTRY_SHEET_NAME)
        return wb

    def _write_dataframe_to_sheet(
        self, ws: Worksheet, df: pl.DataFrame, columns: list[export_config.ColumnConfig]
    ) -> None:
        """Writes a dataframe to a worksheet, using ColumnConfig for headers."""
        db_columns = [c.name for c in columns]
        
        # Ensure all columns exist in the dataframe, adding null ones if missing.
        for col in db_columns:
            if col not in df.columns:
                df = df.with_columns(pl.lit(None, dtype=pl.Utf8).alias(col))
        
        # Write headers using new_name if available
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        for idx, col_config in enumerate(columns, start=1):
            header_name = col_config.new_name or col_config.name
            cell = ws.cell(row=1, column=idx, value=header_name)
            cell.font = header_font
            cell.fill = header_fill

        # Write data rows
        for row_idx, row in enumerate(
            df.select(db_columns).iter_rows(named=True), start=2
        ):
            for col_idx, col_config in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_config.name))

    def _apply_styling_and_validation(self, ws: Worksheet, df: pl.DataFrame) -> None:
        """
        Applies advanced styling, validation, and protection.
        This is a port of the legacy `DataEntrySheet` class logic.
        """
        max_row = df.height
        word_wrap_style = NamedStyle(
            name="wordwrap", alignment=Alignment(wrapText=True)
        )

        for col_idx, col_config in enumerate(
            export_config.DATA_ENTRY_COLUMNS, start=1
        ):
            # Calculate column widths
            col_letter = get_column_letter(col_idx)
            max_width = len(col_config.new_name or col_config.name)
            long_values_count = 0
            for val in df.get_column(col_config.name).drop_nulls():
                val_len = len(str(val))
                if val_len > max_width:
                    max_width = val_len
                if val_len > 40:
                    long_values_count += 1

            # Apply column width / wrap behavior
            if max_width > 40 and (
                (long_values_count > 5) or (long_values_count > max_row - 2)
            ):
                ws.column_dimensions[col_letter].width = 40
                for row_num in range(2, max_row + 2):
                    ws.cell(row=row_num, column=col_idx).style = word_wrap_style
            else:
                ws.column_dimensions[col_letter].width = max(12, min(max_width + 2, 50))

            # Apply data validation
            if col_config.dropdown_options:
                options_str = col_config.dropdown_options
                # If the column is in our enum map, use the enum's labels instead
                if col_config.name in ENUM_MAP:
                    enum_class = ENUM_MAP[col_config.name]
                    options_str = f'"{",".join(str(label) for label in enum_class.labels)}"'
                
                self._add_list_validation(
                    ws, options_str, col_letter, max_row
                )
            
            # Apply hyperlink style
            if col_config.is_url:
                for row_num in range(2, max_row + 2):
                    cell = ws.cell(row=row_num, column=col_idx)
                    if cell.value:
                        cell.hyperlink = cell.value
                        cell.style = "Hyperlink"


        # Apply protection
        unlocked_letters = {
            get_column_letter(i + 1)
            for i, col in enumerate(export_config.DATA_ENTRY_COLUMNS)
            if col.is_editable or col.is_url
        }
        self._protect_sheet_editable(ws, unlocked_letters)

    def _add_list_validation(
        self, ws: Worksheet, options_str: str, col_letter: str, max_row: int
    ):
        """Creates data validation dropdowns using a hidden sheet and named ranges."""
        wb = ws.parent
        list_sheet_name = "_ea_lists"
        if list_sheet_name not in wb.sheetnames:
            list_ws = wb.create_sheet(list_sheet_name)
            list_ws.sheet_state = "hidden"
        else:
            list_ws = wb[list_sheet_name]

        # Use a hash of the options to create a deterministic name for the list
        options_key = options_str.strip('"')
        name_hash = hashlib.md5(options_key.encode("utf-8")).hexdigest()[:10]
        list_name = f"_ea_list_{name_hash}"

        if list_name not in wb.defined_names:
            items = [opt.strip() for opt in options_key.split(",")]
            # Find the first empty column in the hidden sheet
            list_col_idx = 1
            while list_ws.cell(row=1, column=list_col_idx).value is not None:
                list_col_idx += 1
            # Write the options to the hidden sheet
            for row_num, item in enumerate(items, start=1):
                list_ws.cell(row=row_num, column=list_col_idx).value = item
            # Create a named range for the options
            list_col_letter = get_column_letter(list_col_idx)
            ref = f"{quote_sheetname(list_sheet_name)}!${list_col_letter}$1:${list_col_letter}${len(items)}"
            dn = DefinedName(list_name, attr_text=ref)
            wb.defined_names.add(dn)

        # Create and apply the data validation
        dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{max_row + 1}")

    def _create_table(self, ws: Worksheet, max_row: int, style_iter: int):
        """Formats a range as an official Excel Table."""
        max_col_letter = get_column_letter(ws.max_column)
        table = ExcelTable(
            displayName=ws.title.replace(" ", ""),
            ref=f"A1:{max_col_letter}{max_row + 1}",
        )
        style = TableStyleInfo(
            name=f"TableStyleMedium{style_iter}",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    def _protect_sheet_all_locked(self, ws: Worksheet):
        """Protects a sheet by locking all cells."""
        ws.protection.sheet = True
        ws.protection.enable()

    def _protect_sheet_editable(self, ws: Worksheet, editable_letters: set[str]):
        """Protects a sheet, leaving only specified columns editable."""
        # Lock all cells by default
        for row in ws.iter_rows():
            for cell in row:
                cell.protection = cell.protection.copy(locked=True)
        # Unlock editable columns
        for col_letter in editable_letters:
            for cell in ws[col_letter]:
                cell.protection = cell.protection.copy(locked=False)
        # Enable sheet protection
        ws.protection.sheet = True
        ws.protection.enable()

    def _create_empty_workbook_output(self) -> BytesIO:
        """Creates a BytesIO object for an empty workbook."""
        wb = Workbook()
        ws = wb.active
        ws.title = "No Data"
        ws["A1"] = "No data available for export"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output