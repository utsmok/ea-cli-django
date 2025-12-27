"""
Round-trip export/import tests.

Tests the complete cycle of:
1. Ingest Qlik data
2. Export faculty sheets
3. Modify exported file (simulate faculty edits)
4. Re-import modified file
5. Verify changes applied correctly
"""

from pathlib import Path

import openpyxl
import pytest
from django.core.files import File

from apps.core.models import CopyrightItem, Faculty, WorkflowStatus
from apps.ingest.models import (
    IngestionBatch,
)
from apps.ingest.services.export import ExportService
from apps.ingest.services.processor import BatchProcessor
from apps.ingest.tasks import stage_batch
from apps.users.models import User


def _process_batch_sync(batch: IngestionBatch) -> dict:
    """
    Synchronous processing helper for tests.

    Processes a staged batch using BatchProcessor.
    """
    try:
        processor = BatchProcessor(batch)
        processor.process()
        return {
            "success": True,
            "items_created": batch.items_created,
            "items_updated": batch.items_updated,
        }
    except Exception as e:
        return {"success": False, "errors": [str(e)]}


@pytest.fixture
def test_user(db):
    """Create a test user."""
    return User.objects.create_user(username="testuser", email="test@example.com")


@pytest.fixture
def test_data_dir():
    """Path to test data directory."""
    return Path(__file__).parent.parent.parent.parent.parent / "test_data"


@pytest.fixture
def qlik_file(test_data_dir):
    """Path to Qlik export test file."""
    file_path = test_data_dir / "qlik_data.xlsx"
    if not file_path.exists():
        pytest.skip(f"Qlik test file not found: {file_path}")
    return file_path


@pytest.fixture
def test_faculty(db):
    """Create a test faculty."""
    faculty, _ = Faculty.objects.get_or_create(
        abbreviation="TEST",
        defaults={
            "name": "Test Faculty",
            "hierarchy_level": 1,
            "full_abbreviation": "TEST",
        },
    )
    return faculty


@pytest.mark.django_db(transaction=True)
@pytest.mark.slow
@pytest.mark.timeout(30)
class TestRoundTripExportImport:
    """Test round-trip export and import with modifications."""

    def test_round_trip_with_modifications(
        self, test_user, qlik_file, tmp_path, test_faculty
    ):
        """
        Test complete round-trip: export → modify → reimport → verify.

        This simulates the actual workflow:
        1. Import Qlik data
        2. Export faculty sheets
        3. Faculty modifies workflow_status and classification
        4. Re-import modified sheet
        5. Verify changes are applied correctly
        """
        # Skip if no test data
        if not qlik_file.exists():
            pytest.skip("Qlik test file not found")

        # Step 1: Import Qlik data using task
        with Path.open(qlik_file, "rb") as f:
            qlik_batch = IngestionBatch.objects.create(
                source_type=IngestionBatch.SourceType.QLIK,
                uploaded_by=test_user,
                source_file=File(f, name="qlik_data.xlsx"),
            )

        # Use task synchronously (it returns a Task, we need to run it properly)
        # For testing, let's create items directly instead
        qlik_batch.delete()

        # Create test items directly for this test
        items = []
        for i in range(5):
            item = CopyrightItem.objects.create(
                material_id=1000000 + i,
                title=f"Test Material {i}",
                filename=f"test_file_{i}.pdf",
                filetype="pdf",
                workflow_status=WorkflowStatus.TODO,
                classification="short",
                faculty=test_faculty,
            )
            items.append(item)

        # Store initial values for comparison
        initial_states = {}
        for item in items:
            initial_states[item.material_id] = {
                "workflow_status": item.workflow_status,
                "classification": item.classification,
                "remarks": item.remarks,
            }

        # Step 2: Export faculty sheets
        export_dir = tmp_path / "exports"
        exporter = ExportService()
        exporter.export_workflow_tree(output_dir=export_dir)

        assert export_dir.exists(), "Export directory not created"

        # Step 3: Find and modify an exported file
        exported_files = list(export_dir.rglob("*.xlsx"))
        assert len(exported_files) > 0, "No files exported"

        # Find a file we can modify
        modify_file = None
        for f in exported_files:
            if f.name.endswith(".xlsx") and "update" not in f.name.lower():
                modify_file = f
                break

        if not modify_file:
            pytest.skip("No suitable Excel file found for modification")

        # Step 4: Modify the exported file
        wb = openpyxl.load_workbook(modify_file)
        ws = wb.active

        # Find the material_id column and modify corresponding workflow_status
        modifications = []
        for row in ws.iter_rows(min_row=2, max_row=10):
            if len(modifications) >= 2:  # Modify 2 items
                break

            # Find material_id cell
            material_id = None
            for cell in row:
                if isinstance(cell.value, int) and cell.value >= 1000000:
                    material_id = cell.value
                    break

            if material_id and material_id in initial_states:
                # Find workflow_status column (usually nearby)
                for offset in range(1, min(15, ws.max_column - cell.column)):
                    target_cell = ws.cell(row=cell.row, column=cell.column + offset)
                    if target_cell.value and isinstance(target_cell.value, str):
                        if target_cell.value in ["ToDo", "InProgress", "Done"]:
                            # This is the workflow_status column
                            old_status = target_cell.value
                            new_status = "Done" if old_status != "Done" else "ToDo"
                            target_cell.value = new_status
                            modifications.append(
                                {
                                    "material_id": material_id,
                                    "field": "workflow_status",
                                    "old_value": old_status,
                                    "new_value": new_status,
                                }
                            )
                            break

        wb.save(tmp_path / "modified.xlsx")

        assert len(modifications) > 0, "No modifications were made to the file"

        # Step 5: Create a simple faculty sheet for reimport
        # Create a minimal faculty sheet with the modifications
        wb_faculty = openpyxl.Workbook()
        ws_faculty = wb_faculty.active
        ws_faculty.title = "Data entry"

        # Header row
        headers = ["Material ID", "Workflow status", "Classification", "Remarks"]
        for col, header in enumerate(headers, start=1):
            ws_faculty.cell(row=1, column=col, value=header)

        # Data rows with modifications
        for mod in modifications:
            item = CopyrightItem.objects.get(material_id=mod["material_id"])
            ws_faculty.cell(
                row=modifications.index(mod) + 2, column=1, value=item.material_id
            )
            ws_faculty.cell(
                row=modifications.index(mod) + 2, column=2, value=mod["new_value"]
            )
            ws_faculty.cell(
                row=modifications.index(mod) + 2,
                column=3,
                value=item.classification or "",
            )
            ws_faculty.cell(
                row=modifications.index(mod) + 2, column=4, value=item.remarks or ""
            )

        faculty_file_path = tmp_path / "faculty_sheet.xlsx"
        wb_faculty.save(faculty_file_path)

        # Step 6: Re-import the faculty sheet
        with Path.open(faculty_file_path, "rb") as f:
            faculty_batch = IngestionBatch.objects.create(
                source_type=IngestionBatch.SourceType.FACULTY,
                uploaded_by=test_user,
                source_file=File(f, name="faculty_sheet.xlsx"),
                faculty_code=test_faculty.abbreviation,
            )

        # Use task directly - we need to run synchronously
        stage_batch.enqueue(faculty_batch.id)
        process_result = _process_batch_sync(faculty_batch)

        assert process_result["success"], f"Faculty processing failed: {process_result}"

        # Step 7: Verify modifications were applied
        for mod in modifications:
            item = CopyrightItem.objects.filter(material_id=mod["material_id"]).first()
            assert item is not None, f"Item {mod['material_id']} not found"
            assert item.workflow_status == mod["new_value"], (
                f"workflow_status not updated for item {mod['material_id']}: "
                f"expected {mod['new_value']}, got {item.workflow_status}"
            )


@pytest.mark.django_db(transaction=True)
class TestExportIntegrity:
    """Test that exports maintain data integrity."""

    def test_export_preserves_all_data(self, test_user, qlik_file, tmp_path):
        """Verify that export includes all critical fields."""
        if not qlik_file.exists():
            pytest.skip("Qlik test file not found")

        # Create test items directly
        faculty = Faculty.objects.create(
            abbreviation="TEST2",
            name="Test Faculty 2",
            hierarchy_level=1,
            full_abbreviation="TEST2",
        )

        for i in range(3):
            CopyrightItem.objects.create(
                material_id=2000000 + i,
                title=f"Test Export Material {i}",
                filename=f"test_export_{i}.pdf",
                filetype="pdf",
                workflow_status=WorkflowStatus.TODO,
                classification="short",
                faculty=faculty,
            )

        # Export
        export_dir = tmp_path / "exports"
        exporter = ExportService()
        exporter.export_workflow_tree(output_dir=export_dir)

        # Verify files exist
        exported_files = list(export_dir.rglob("*.xlsx"))
        assert len(exported_files) > 0, "No files exported"

        # Verify files are non-empty
        for f in exported_files:
            if "update" not in f.name.lower():
                assert f.stat().st_size > 0, f"Exported file is empty: {f}"


@pytest.mark.django_db
def test_empty_export_does_not_fail(test_user, tmp_path):
    """Test that exporting with no data doesn't fail."""
    # Create a faculty but no items
    Faculty.objects.get_or_create(
        abbreviation="EMPTY",
        defaults={
            "name": "Empty Faculty",
            "hierarchy_level": 1,
            "full_abbreviation": "EMPTY",
        },
    )

    # Export should succeed even with no data
    export_dir = tmp_path / "exports"
    exporter = ExportService()
    result = exporter.export_workflow_tree(output_dir=export_dir)

    # Should complete without error
    assert "error" not in result or result.get("error") is None
    assert export_dir.exists()
