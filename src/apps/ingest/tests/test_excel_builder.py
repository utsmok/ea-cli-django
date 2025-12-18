from io import BytesIO

import pytest
from openpyxl import load_workbook

from apps.core.models import CopyrightItem, Faculty
from apps.ingest.services.excel_builder import ExcelBuilder


@pytest.mark.django_db
def test_excel_builder_creates_two_sheet_workbook(tmp_path):
    """Test that ExcelBuilder creates a workbook with 'Complete data' and 'Data entry' sheets."""
    faculty = Faculty.objects.create(
        hierarchy_level=1,
        name="Electrical Engineering",
        abbreviation="EEMCS",
        full_abbreviation="EEMCS",
    )

    CopyrightItem.objects.create(
        material_id=123,
        title="Test Item",
        faculty=faculty,
    )

    # Fetch data for EEMCS
    from apps.ingest.services.export import ExportService

    service = ExportService(faculty_abbr="EEMCS")
    df = service._fetch_faculty_dataframe("EEMCS")

    # Build workbook
    builder = ExcelBuilder()
    wb = builder.build_workbook_for_dataframe(df)

    # Save to BytesIO for testing
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Reload and verify
    wb = load_workbook(output)

    assert ExcelBuilder.COMPLETE_SHEET_NAME in wb.sheetnames
    assert ExcelBuilder.DATA_ENTRY_SHEET_NAME in wb.sheetnames

    # Check that complete data sheet has all columns
    ws_complete = wb[ExcelBuilder.COMPLETE_SHEET_NAME]
    complete_headers = [cell.value for cell in ws_complete[1]]
    assert len(complete_headers) > 0

    # Check that data entry sheet has expected columns
    ws_entry = wb[ExcelBuilder.DATA_ENTRY_SHEET_NAME]
    entry_headers = [cell.value for cell in ws_entry[1]]
    assert len(entry_headers) > 0

    # Verify that data entry sheet is the active sheet
    assert wb.active.title == ExcelBuilder.DATA_ENTRY_SHEET_NAME


@pytest.mark.django_db
def test_excel_builder_with_multiple_items():
    """Test that ExcelBuilder handles multiple items correctly."""
    faculty = Faculty.objects.create(
        hierarchy_level=1,
        name="Behavioural Sciences",
        abbreviation="BMS",
        full_abbreviation="BMS",
    )

    # Create multiple items
    for idx in range(3):
        CopyrightItem.objects.create(
            material_id=idx + 1,
            title=f"Item {idx + 1}",
            faculty=faculty,
            count_students_registered=10,
        )

    from apps.ingest.services.export import ExportService

    service = ExportService(faculty_abbr="BMS")
    df = service._fetch_faculty_dataframe("BMS")

    builder = ExcelBuilder()
    wb = builder.build_workbook_for_dataframe(df)

    # Check that both sheets have data
    ws_complete = wb[ExcelBuilder.COMPLETE_SHEET_NAME]
    ws_entry = wb[ExcelBuilder.DATA_ENTRY_SHEET_NAME]

    # Check row counts (1 header + 3 data rows)
    assert ws_complete.max_row == 4
    assert ws_entry.max_row == 4


@pytest.mark.django_db
def test_excel_builder_column_validation():
    """Test that data entry columns have dropdowns for editable fields."""
    faculty = Faculty.objects.create(
        hierarchy_level=1,
        name="Test Faculty",
        abbreviation="TEST",
        full_abbreviation="TEST",
    )

    CopyrightItem.objects.create(
        material_id=999,
        title="Test Item",
        faculty=faculty,
    )

    from apps.ingest.services.export import ExportService

    service = ExportService(faculty_abbr="TEST")
    df = service._fetch_faculty_dataframe("TEST")

    builder = ExcelBuilder()
    wb = builder.build_workbook_for_dataframe(df)

    ws_entry = wb[ExcelBuilder.DATA_ENTRY_SHEET_NAME]

    # Verify that data validations exist
    assert len(ws_entry.data_validations.dataValidation) > 0

    # Verify hidden sheet for dropdown options exists
    assert "_ea_lists" in wb.sheetnames
    list_sheet = wb["_ea_lists"]
    assert list_sheet.sheet_state == "hidden"
